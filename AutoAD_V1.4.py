 #coding=utf-8
 #-*- encoding: utf-8 -*-

#---------------------------------------------------一. 导入开源模块-------------------------------------------------
#python2的py文件里面写中文，则必须要添加第1行的声明文件编码的注释
import os  # 获取文件路径需要这个模块
import openpyxl  # 写xlsx文件需要的模块
from openpyxl.styles import PatternFill
import math   #因为广告算法中需要用到e的相关计算，所以因为数学库
import matplotlib.pyplot as plt

#全局变量定义和声明
ACOS_TARGET = 0.36
LOW_CLICkS = 5
UP_BID = 3
DOWN_BID = 0.2
rBidCpc = 1.2   # 因为Bid可能有人会在后台直接修改，所以算新bid的时候用的是旧cpc而不是旧bid。但是cpc往往比bid小一点，所以要乘以一个放大系数。
Price = 12.6
Clicks_Base = 14

#----------------------------------------------三. 功能实现，之计算每个excel里BidNew-------------------------------------------------
# 函数，将文本中的百分数转换为小数
def AD_Perc2Delc(percent):
    # print(percent[0:-1])
    p_float = float(percent[0:-1])/100
    p_float_2 = round(p_float, 2)
    return p_float_2

# 函数，根据输入转换fileIndex
def AD_ExchangeInput2FileIndex(inputfile):
    fileIndex = 0

    if (inputfile == '14'):
        fileIndex = 0
    elif(inputfile == '30'):
        fileIndex = 1
    elif(inputfile == '60'):
        fileIndex = 2
    return fileIndex

# 函数，从excel读出来的源数据是字符串格式，下面格式转换为可以计算的数字格式
def AD_FmtSrcData(src_head, src_ws):
    AcosStr = src_ws.cell(row=src_head, column=35+1).value
    # 将Acos的字符串百分数转换为小数
    Acos = AD_Perc2Delc(AcosStr)
    Clicks = int(src_ws.cell(row=src_head, column=28+1).value)
    Unit = int(src_ws.cell(row=src_head, column=33+1).value)
    CPC = float(src_ws.cell(row=src_head, column=36+1).value)
    BidOld = float(src_ws.cell(row=src_head, column=21).value)
    return BidOld, Acos, Clicks, Unit, CPC

# 广告算法，根据Acos计算BidNew        
def AD_BidNewAlg(BidOld, Acos, Clicks, Unit, CPC):
    BidNew = 0

    if Acos == 0:
        if Clicks <= LOW_CLICkS:
            BidNew = BidOld
        else :
            Cr_Est = 0.5 * (1 / (Clicks + 1))
            Acos = CPC / (Cr_Est * Price)
            x = ACOS_TARGET / Acos
            if x > 1:
                #y =−(1/𝑒) ^ (x−1)+2
                y = (-1 / (math.exp(x-1))) + 2                        
            else: 
                # y = 𝑒 ^ (x−1) 
                y = math.exp(x-1)
            BidNew = rBidCpc * CPC * y
    if Acos > 0:
        if Clicks <= LOW_CLICkS:
            Cr_Est = 0.5 * (1/(Clicks))
            Acos = CPC / (Cr_Est * Price)
        x = ACOS_TARGET / Acos
        if x > 1:
            #y =−(1/𝑒) ^ (x−1)+2
            y = (-1 / (math.exp(x-1))) + 2 
        else:  
            # y = 𝑒 ^ (x−1)
            y = math.exp(x-1)
        BidNew = rBidCpc * CPC * y
        # 数据保留2位小数    
    BidNew = round(BidNew,2)
    return BidNew

# 函数，BidNew处理模块，处理每行
def AD_DataProc4EachLine(src_head, search_flag, src_ws, print_log):
    # 筛选Product Targeting和Keyword的行
    if search_flag == 'Product Targeting' or search_flag == 'Keyword':            
        BidOld = src_ws.cell(row=src_head, column=21).value
        Keyword_Id = src_ws.cell(row=src_head, column=8).value
        Product_Id = src_ws.cell(row=src_head, column=9).value                               
        # 如果Bid为空则标记红色出来并退出
        if BidOld =="" or BidOld is None:
            fille = PatternFill('solid',fgColor="00FF0000") #标记为红色
            src_ws.cell(row=src_head, column=2).fill = fille
            print("line ", src_head,"BidOld is None, ", "Keyword_Id:", Keyword_Id, "Product_Id:", Product_Id, file = print_log)
            return
        else:
            # 获得格式化后的源数据
            BidOld, Acos, Clicks, Unit, CPC = AD_FmtSrcData(src_head, src_ws)
            src_ws.cell(row=src_head, column=20).value = BidOld
            # 打印检查获取数据是否正确
            print('\nline', src_head,' < Keyword_Id',Keyword_Id,'> <','Product_Id',Product_Id,'>', file = print_log)
            print(' orig data: ','BidOld',BidOld, 'Acos',Acos, 'Clicks',Clicks, 'Unit',Unit, 'CPC', CPC, file = print_log)
            # 计算BidNew
            BidNew = AD_BidNewAlg(BidOld, Acos, Clicks, Unit, CPC)                        
        # 打印最终结果
        print('line', src_head,' < Keyword_Id',Keyword_Id,'> <','Product_Id',Product_Id,'> <','upper bound',round(UP_BID*BidOld,2),'> <','lower bound',round(DOWN_BID*BidOld,2),'>', file = print_log)
        print('final data: ',"BidOld",BidOld, "BidNew", BidNew, "Acos",Acos, "Clicks",Clicks, "Unit",Unit, "CPC", CPC, file = print_log)
        print('\n', file = print_log)
        return BidNew
    else:
        print("line ", src_head,"is not 'Product Targeting' or 'Keyword', ", "search_flag:", search_flag, file=print_log)
        return

# 函数，BidNew处理模块，处理每个文件
def AD_BidNewProc(src_head, src_ws, print_log):
    #---------------------------------- 2.1 打开文件---------------------------------------
    src_ws.insert_cols(20, 1)
    src_ws.cell(row=1, column=20).value = 'BidNew'
    # 检查行数，以及下面的列号取的对不对
    print("src_ws.max_row ", src_ws.max_row)
    if ("Bid" == src_ws.cell(row=1, column=21).value) and\
        ("Keyword Id (Read only)" == src_ws.cell(row=1, column=8).value) and\
        ("Product Targeting Id (Read only)" == src_ws.cell(row=1, column=9).value) and\
        ("Clicks" == src_ws.cell(row=1, column=28+1).value) and\
        ("Units" == src_ws.cell(row=1, column=33+1).value) and\
        ("Acos" == src_ws.cell(row=1, column=35+1).value) and\
        ("CPC" == src_ws.cell(row=1, column=36+1).value):
        print("列号计算正确")
    else:
        print("Bid is ", src_ws.cell(row=1, column=21).value)
        print("Keyword_Id ", src_ws.cell(row=1, column=8).value)
        print("Product_Id ", src_ws.cell(row=1, column=9).value)
        print("Clicks ", src_ws.cell(row=1, column=28+1).value)
        print("Unit ", src_ws.cell(row=1, column=33+1).value)
        print("AcosStr ", src_ws.cell(row=1, column=35+1).value)
        print("CPC ", src_ws.cell(row=1, column=36+1).value)

    # 循环处理每行
    for row in range(2, src_ws.max_row+1):
        # src_head标记for循环到的当前行数
        src_head = src_head + 1
        search_flag = src_ws.cell(row=src_head, column=2).value
        BidNew = AD_DataProc4EachLine(src_head, search_flag, src_ws, print_log)
        src_ws.cell(row=src_head, column=20).value = BidNew

#--------------------------------------四. 功能实现之计算14天excel里BidAvg---------------------------------
print('\n-----***-----\n')
print('功能实现之计算14天excel里BidAvg\n')
print('\n-----***-----\n')

# 广告算法，根据Click和BidNew计算BidAvg   
def AD_BidAvgAlg(Clicks_14, Clicks_30, Clicks_60, BidNew_14, BidNew_30, BidNew_60, CPC_14):
    #timeCoe14\timeCoe30\timeCoe60是时间权重
    if Clicks_14 <= 20:
        timeCoe14 = 0.5
        timeCoe30 = 0.3
        timeCoe60 = 0.2 
    elif Clicks_14 > 20 and Clicks_14 <= 50:
        timeCoe14 = 0.7
        timeCoe30 = 0.2
        timeCoe60 = 0.1
    elif Clicks_14 > 50 and Clicks_14 <= 100:
        timeCoe14 = 0.8
        timeCoe30 = 0.15
        timeCoe60 = 0.05
    else:
        timeCoe14 = 0.9
        timeCoe30 = 0.1
        timeCoe60 = 0
    
    weight14 = timeCoe14 * Clicks_14
    weight30 = timeCoe30 * Clicks_30
    weight60 = timeCoe60 * Clicks_60
    weightSum = weight14 + weight30 + weight60
    
    BidAvg = float(BidNew_14) * weight14 / weightSum+ float(BidNew_30) * weight30 / weightSum + float(BidNew_60) * weight60 / weightSum

    if BidAvg > CPC_14:
        if Clicks_14 > Clicks_Base:
            fille = PatternFill('solid',fgColor="00008000") #深绿色       
        else:
            fille = PatternFill('solid',fgColor="00CCFFCC") #浅绿色
    else:
        if Clicks_14 > Clicks_Base:
            fille = PatternFill('solid',fgColor="00FF6600") #深橙色       
        else:
            fille = PatternFill('solid',fgColor="00FFCC99") #浅橙色    

    return BidAvg, fille

# 函数，BidAvg处理模块，处理每个行
def AD_BidAvgProc(src_head, search_flag, src_ws_14, src_ws_30, src_ws_60, BidChgList, BidOldList, BidAvgList, print_log):
    if search_flag == 'Product Targeting' or search_flag == 'Keyword':            
        BidOld_14 = src_ws_14.cell(row=src_head, column=22).value
        Keyword_Id = src_ws_14.cell(row=src_head, column=8).value
        Product_Id = src_ws_14.cell(row=src_head, column=9).value
        # 如果Bid为空则退出
        if BidOld_14 =="" or BidOld_14 is None:
            print("line ", src_head,"BidOld_14 is None, ", "Keyword_Id:", Keyword_Id, "Product_Id:", Product_Id, file = print_log)
            return
        else:
            Clicks_14 = int(src_ws_14.cell(row=src_head, column=30).value)
            BidNew_14 = src_ws_14.cell(row=src_head, column=21).value
            # print('Keyword_Id', Keyword_Id,'Product_Id',Product_Id,'Clicks_14',Clicks_14, 'BidOld_14',BidOld_14,'BidNew_14',BidNew_14)
            if Keyword_Id is not None:
                keyId_column = src_ws_30['H']
                head = 0
                for cell in range(2,len(keyId_column)+1):
                    head=cell+1
                    if keyId_column[head].value == Keyword_Id:
                        flag = head+1
                        # print('src_ws_30',head, flag)
                        Clicks_30 = int(src_ws_30.cell(row=flag, column=29).value)
                        BidNew_30 = src_ws_30.cell(row=flag, column=20).value
                        # print('Clicks_30',Clicks_30,'BidNew_30',BidNew_30)
                        break
                keyId_column = src_ws_60['H']
                head = 0
                for cell in range(2,len(keyId_column)+1):
                    head=cell+1
                    if keyId_column[head].value == Keyword_Id:
                        flag=head+1
                        # print('src_ws_60',head, flag)
                        Clicks_60 = int(src_ws_60.cell(row=flag, column=29).value)
                        BidNew_60 = src_ws_60.cell(row=flag, column=20).value
                        # print('Clicks_60',Clicks_60,'BidNew_60',BidNew_60)
                        break
            if Product_Id is not None:
                productId_column = src_ws_30['I']
                for cell in range(2,len(productId_column)):
                    head=cell+1
                    if productId_column[cell].value == Product_Id:
                        flag=head
                        # print('src_ws_30',head, flag)
                        Clicks_30 = int(src_ws_30.cell(row=flag, column=29).value)
                        BidNew_30 = src_ws_30.cell(row=flag, column=20).value
                        # print('Clicks_30',Clicks_30,'BidNew_30',BidNew_30)
                        break
                productId_column = src_ws_60['I']
                for cell in range(2,len(productId_column)):
                    head=cell+1
                    if productId_column[cell].value == Product_Id:
                        flag=head
                        # print('src_ws_60',head, flag)
                        Clicks_60 = int(src_ws_60.cell(row=flag, column=29).value)
                        BidNew_60 = src_ws_60.cell(row=flag, column=20).value
                        # print('Clicks_60',Clicks_60,'BidNew_60',BidNew_60)
                        break

            ClicksSum = Clicks_14 + Clicks_30 + Clicks_60
            if BidNew_14 is None or BidNew_30 is None or BidNew_60 is None:
                print('\n<Keyword_Id',Keyword_Id,'> <','Product_Id',Product_Id,'>', file=print_log)
                print('BidNew_14',BidNew_14,'BidNew_30',BidNew_30,'BidNew_60',BidNew_60, file=print_log)
                return
            if ClicksSum == 0:
                print('<Keyword_Id',Keyword_Id,'> <','Product_Id',Product_Id,'>', file=print_log)
                print('ClicksSum is 0\n', file=print_log)
                return

            print("\nline ", src_head, '<Keyword_Id',Keyword_Id,'> <','Product_Id',Product_Id,'>', file = print_log)
            print('ClicksSum',ClicksSum,' ','Clicks_14',Clicks_14,' ','Clicks_30',Clicks_30,' ','Clicks_60',Clicks_60, file = print_log)
            print('BidNew_14',BidNew_14, 'BidNew_30',BidNew_30,'BidNew_60',BidNew_60, file = print_log)

            CPC_14 = float(src_ws_14.cell(row=src_head, column=38).value)
            BidAvg, fille = AD_BidAvgAlg(Clicks_14, Clicks_30, Clicks_60, BidNew_14, BidNew_30, BidNew_60, CPC_14)

            BidChgList[src_head] = BidAvg - CPC_14
            BidOldList[src_head] = src_ws_14.cell(row=src_head,column=22)
            BidAvgList[src_head] = BidAvg
            src_ws_14.cell(row=src_head,column=2).fill = fille
            src_ws_14.cell(row=src_head, column=20).value = round(BidAvg,2)
            src_ws_14.cell(row=src_head, column=3).value = 'Update'
            print('final: BidAvg',round(BidAvg,2), file=print_log)
            return
    else:
        print("line ", src_head,"is not 'Product Targeting' or 'Keyword', ", "search_flag:", search_flag, file = print_log)
        return

# 函数，画图模块
def AD_PlotResult(row, BidOldList, BidAvgList, BidChgList):

    list = [i for i in range(row)]
    plt.title('每行商品BigAvg相比CPC调整情况')
    # 设置坐标轴
    plt.xlabel('行号')
    plt.ylabel('BidAvg-CPC_14')
    # 设置坐标轴范围
    plt.xticks([0,row])
    plt.yticks([-1,6])
    # 画两条虚线
    plt.hlines(1,0,row, colors='r', linestyle='--')
    plt.hlines(-1,0,row, colors='g',linestyle='--')
    plt.scatter(x= list, y=BidOldList, marker='X', c=BidOldList, cmap='coolwarm')
    plt.scatter(x= list, y=BidAvgList, marker='o', c=BidAvgList, cmap='coolwarm')
    plt.scatter(x= list, y=BidChgList, marker='*', c=BidChgList, cmap='coolwarm')
    plt.savefig('./Bid变化图.png')
    plt.show()

#------------------------------------------------------- 3 Main函数-------------------------------------------------------

if __name__ == '__main__':    

    # #--------------------------------一. 获得待处理的目标文件目录------------------------
    # # 获取python脚本文件所在绝对路径
    # pyFilePath = os.path.abspath(__file__)    
    # # 获取python脚本文件所在目录
    # pyFileDir = os.path.dirname(pyFilePath)   
    # print(pyFileDir, pyFilePath)  
    # # 当有多个文件夹下数据需要处理时，可通过此命令选择待处理文件夹
    # processingDir = input("please enter your target folder: ")
    # print('-----***-----\n')
    # filesDir = pyFileDir+'./'+processingDir
    # # 打印所有目标文件
    # filesArray = os.listdir(processingDir)    
    # print('All the files are:\n', filesArray)
    # print('\n-----***-----\n')
    # # 创建输出文件夹及目录
    # outputDir = filesDir                      
    # print('输出文件目录:', outputDir)
    # # 获取输出目录下所有的excel表
    # file_list = os.listdir(outputDir)         
    # print('-----***-----\n')

    # #---------------------------二. 功能实现, 之计算每个excel里BidNew---------------------
    # with open(outputDir + "_calBidNewlog.txt","w") as print_log:
    #     for i in range(3):
    #         inputfile = input("input 14 or 30 or 60: ")
    #         fileIndex = AD_ExchangeInput2FileIndex(inputfile)
    #         file_path = outputDir + '/' + file_list[fileIndex]
    #         print(file_path, '\n')
    #         src_head = 1
    #         # 打开待处理文件
    #         src_wb = openpyxl.load_workbook(file_path)
    #         src_ws = src_wb['Sponsored Products Campaigns']
    #         # 计算每个文件里所有BidNew
    #         AD_BidNewProc(src_head, src_ws, print_log)
    #         # 保存文件
    #         src_wb.save(file_path)
    #         print(file_path ," 计算完成\n")
    #     print('-------------all file cal BidNew finished-----------\n')
    #     print('\n----------------------end------------------------\n', file=print_log)
    # print_log.close

    # #---------------------------三. 功能实现, 之计算14天excel里BidAvg----------------------
    # src_wb_14= openpyxl.load_workbook(outputDir + '/' +'14.xlsx')
    # src_wb_30= openpyxl.load_workbook(outputDir + '/' +'30.xlsx')
    # src_wb_60= openpyxl.load_workbook(outputDir + '/' +'60.xlsx')        
    # src_ws_14 = src_wb_14['Sponsored Products Campaigns']
    # src_ws_30 = src_wb_30['Sponsored Products Campaigns']
    # src_ws_60 = src_wb_60['Sponsored Products Campaigns']
    # # 插入BidAvg的列
    # src_ws_14.insert_cols(20,1)
    # src_ws_14.cell(row=1, column=20).value ='BidAvg'
    
    BidChgList = [5 for i in range(100)]
    BidAvgList = [5 for i in range(100)]
    BidOldList = [5 for i in range(100)]
    # file_path = outputDir + '/' +'总结.xlsx'
    # #记录打印日志
    # with open(outputDir + "_calBidAvglog.txt","w") as print_log:
    #     for row in range(2, src_ws_14.max_row+1):
    #         src_head=src_head+1
    #         # 计算所有BidAvg
    #         search_flag = src_ws_14.cell(row=src_head, column=2).value
    #         AD_BidAvgProc(src_head, search_flag, src_ws_14, src_ws_30, src_ws_60, BidChgList, BidOldList, BidAvgList, print_log)
       
    #     print('\n----------------------end------------------------\n', file=print_log)
    # print_log.close

    # src_wb_14.save(file_path)
    # print('-----***----\n')
    
    AD_PlotResult(100, BidOldList, BidAvgList, BidChgList)



