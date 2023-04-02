# coding=utf-8
# -*- encoding: utf-8 -*-

# ---------------------------------------------------一. 导入开源模块-------------------------------------------------
# python2的py文件里面写中文，则必须要添加第1行的声明文件编码的注释
import os  # 获取文件路径需要这个模块
import openpyxl  # 写xlsx文件需要的模块
import xml.dom.minidom  # 读写xml文件需要的模块
from pandas import read_csv  # 读csv文件需要的模块
from pandas import DataFrame  # 存储xml信息到xlsx文件需要的模块
# DataFrame是Pandas中的一个表结构的数据结构，包括三部分信息，表头（列的名称），表的内容（二维矩阵），索引（每行一个唯一的标记）。
import pandas as pd
import shutil  # 复制文件需要的模块
import warnings  # 屏蔽一些不影响功能的告警

warnings.filterwarnings('ignore')


# ----------------------------------------二. 功能实现之获得转换格式后的临时汇总报告-----------------------------------
# ---------- 1.1.2 创建输出文件夹及目录
def mkdir(path):
    folder = os.path.exists(path)
    # 判断是否存在文件夹如果不存在则创建为文件夹
    if not folder:
        os.mkdir(path)
    else:
        print("The folder exists already!")


# --------1.2 获取各种类型的待处理文件
def business_getfile(filesArray):
    csvfileName, csvInput, txtfileName, txtInput, xlsxfileName, xlsxInput, xmlfileName, xmlInput = "", "", "", "", "", "", "", ""
    xmlfileName = ''
    for file in filesArray:
        # find函数返回字符串中查找字符所在位置
        pos = file.find('.')
        iFormat = file[pos + 1:]
        if iFormat == 'csv':
            # 获得不带后缀的文件名
            csvfileName = file[:pos]
            # 保存输入文件的变量
            csvInput = file
        elif iFormat == 'txt':
            txtfileName = file[:pos]
            txtInput = file
        elif iFormat == 'xml':
            xmlfileName = file[:pos]
            xmlInput = file
        elif iFormat == 'xlsx':
            xlsxfileName = file[:pos]
            xlsxInput = file
        else:
            otherfileName = os.path.splitext(file)[0]
            otherInput = file
    return csvfileName, csvInput, txtfileName, txtInput, xlsxfileName, xlsxInput, xmlfileName, xmlInput


# 将文本中的百分数转换为小数
def perc2delc(percent):
    p_float = percent.str.strip("%").astype(float) / 100
    p_float_2 = p_float.round(decimals=4)
    return p_float_2


# 去掉文本中的欧元符号
def cur2delc(percent):
    value = percent.str.replace(',', '').str.strip("€").astype(float)
    return value

def business_csv2xlsx(filesDir, outputDir, csvfileName, csvInput):
    # --------2.1 csv转xlsx
    inputFile = open(filesDir + '/' + csvInput, 'rb')
    # 读取以分号为分隔符的csv文件,sep作用为指定分隔符，默认在Windows系统系分隔符为逗号，index_col=0的作用是去除excel中的第一列行号，thousands是为了识别千位数
    data_csv = pd.read_csv(inputFile, encoding='utf-8', sep=',', index_col=0, thousands=',')
    percent = data_csv['会话百分比 – 总计']
    data_csv['会话百分比 – 总计'] = perc2delc(data_csv['会话百分比 – 总计'])
    data_csv['页面浏览量百分比 – 总计'] = perc2delc(data_csv['页面浏览量百分比 – 总计'])
    data_csv['推荐报价（购买按钮）百分比'] = perc2delc(data_csv['推荐报价（购买按钮）百分比'])
    data_csv['商品会话百分比'] = perc2delc(data_csv['商品会话百分比'])
    data_csv['商品会话百分比 - B2B'] = perc2delc(data_csv['商品会话百分比 - B2B'])
    data_csv['已订购商品销售额'] = cur2delc(data_csv['已订购商品销售额'])
    data_csv['已订购商品销售额 - B2B'] = cur2delc(data_csv['已订购商品销售额 - B2B'])
    # 数据输出
    data_csv.to_excel(outputDir + '1' + csvfileName + '.xlsx')
    print('csv文件格式转换完毕')
    return data_csv


def getMonth(sheet):
    # 算法解释：如果第一行和中间一行的月份相同，那么这一周过半数据属于当前月，否则，这一周过半数据属于下一月，那么月份就取下一月
    start_tmp = str(sheet.cell(row=2, column=1).value).split('-')
    start_Month = start_tmp[1]
    end_tmp = str(sheet.cell(row=sheet.max_row - 1, column=1).value).split('-')
    end_Month = end_tmp[1]
    middle_tmp = str(sheet.cell(row=((sheet.max_row - 1) // 2), column=1).value).split('-')
    middle_Month = middle_tmp[1]
    if start_Month == middle_Month:
        theMonth = start_Month
    else:
        theMonth = end_Month
    return theMonth


def business_txt2xlsx(outputDir, filesDir, txtfileName, txtInput):
    # --------2.2 txt转xlsx
    # 数据输出
    outPutFile = outputDir + '3' + txtfileName + '.xlsx'
    filename = filesDir + '/' + txtInput
    xlsxname = outPutFile

    theYear, theMonth = 0, 0
    try:
        file = open(filename, 'rb')
        xlsx = openpyxl.Workbook()
        # 生成excel的方法，声明excel
        sheet = xlsx.active
        # 获取文本文件所有行
        lines = file.readlines()
        x = 0  # 行号
        for line in lines:
            # 按行循环，读取文本文件
            x = x + 1
            line = str(line, 'utf-8')
            # print(line)
            if not line:
                break  # 如果没有内容，则退出循环
            for col in range(len(line.split('\t'))):
                item = line.split('\t')[col]
                if (x > 1 and col == 6):
                    sheet.cell(row=x, column=col + 1).value = float(item)
                    # print(item)
                else:
                    sheet.cell(row=x, column=col + 1).value = item

        file.close()
        sheet.cell(row=x, column=col + 1).value = item
        theYear = str(sheet.cell(row=2, column=1).value)[:4]
        theMonth = getMonth(sheet)
        #print("theYear", theYear, "theMonth", theMonth)
        xlsx.save(xlsxname)  # 保存xlsx文件
    except Exception as e:
        print(e)
    print('txt文件格式转换完毕')
    return theYear, theMonth


def business_xml2xlsx(xmlfileName, filesDir, xmlInput, pyFileDir, outputDir, xlsxInput, xlsxfileName):
    # --------2.3 xml转xlsx
    if xmlfileName == '':
        shutil.copy(pyFileDir + './自发货退货_空表.xlsx', outputDir + '4自发货退货_空表.xlsx')
    else:
        # 打开xml文档
        dom = xml.dom.minidom.parse(filesDir + '/' + xmlInput)

        # data是一个类似字典的数据结构，索引+值
        data = {
            'DocumentVersion': [dom.getElementsByTagName('DocumentVersion')[0].firstChild.data],
            'MessageType': [dom.getElementsByTagName('MessageType')[0].firstChild.data],
            'item_name': [dom.getElementsByTagName('item_name')[0].firstChild.data],
            'asin': [dom.getElementsByTagName('asin')[0].firstChild.data],
            'return_reason_code': [dom.getElementsByTagName('return_reason_code')[0].firstChild.data],
            'merchant_sku': [dom.getElementsByTagName('merchant_sku')[0].firstChild.data],
            'in_policy': [dom.getElementsByTagName('in_policy')[0].firstChild.data],
            'return_quantity': [dom.getElementsByTagName('return_quantity')[0].firstChild.data],
            'resolution': [dom.getElementsByTagName('resolution')[0].firstChild.data],
            'category': [dom.getElementsByTagName('category')[0].firstChild.data],
            'refund_amount': [float(dom.getElementsByTagName('refund_amount')[0].firstChild.data)],
            'order_id': [dom.getElementsByTagName('order_id')[0].firstChild.data],
            'order_date': [dom.getElementsByTagName('order_date')[0].firstChild.data],
            'amazon_rma_id': [dom.getElementsByTagName('amazon_rma_id')[0].firstChild.data],
            'return_request_date': [dom.getElementsByTagName('return_request_date')[0].firstChild.data],
            'return_request_status': [dom.getElementsByTagName('return_request_status')[0].firstChild.data],
            'a_to_z_claim': [dom.getElementsByTagName('a_to_z_claim')[0].firstChild.data],
            'is_prime': [dom.getElementsByTagName('is_prime')[0].firstChild.data],
            'label_cost': [float(dom.getElementsByTagName('label_cost')[0].firstChild.data)],
            'label_type': [dom.getElementsByTagName('label_type')[0].firstChild.data],
            'label_to_be_paid_by': [dom.getElementsByTagName('label_to_be_paid_by')[0].firstChild.data],
            'return_type': [dom.getElementsByTagName('return_type')[0].firstChild.data],
            'order_amount': [float(dom.getElementsByTagName('order_amount')[0].firstChild.data)],
            'order_quantity': [dom.getElementsByTagName('order_quantity')[0].firstChild.data],
        }
        # 保存为DataFrame数据结构
        df = DataFrame(data)
        # 数据输出
        xmlfile = outputDir + '4' + xmlfileName + '.xlsx'
        df.to_excel(xmlfile)
        wb = openpyxl.load_workbook(xmlfile)
        ws = wb.active
        ws.delete_cols(1)
        wb.save(xmlfile)
    print('xml文件格式转换完毕')
    # --------2.4 将原来已有的xlsx文件复制到output文件夹

    print('所有文件格式转换完毕')
    print('-----***-----\n')
    shutil.copy(filesDir + './' + xlsxInput, outputDir + '2' + xlsxfileName + '.xlsx')


def business_combileXlsx(outputDir, filesDir, processingDir):
    # --------3.1 获取输出目录下所有的excel表
    file_list = os.listdir(outputDir)
    print('The processed files are:\n', file_list)
    # --------3.2 创建一个新的excel表，综合报告
    result = openpyxl.Workbook()
    # --------3.3 循环依次读取待合并的工作表
    for file in file_list:
        file_path = outputDir + file
        # print(file_path)
        # 获得待合并表的表名
        sheet_name = file.split('/')[-1].split('.')[0]
        # print(sheet_name)
        old_wb = openpyxl.load_workbook(file_path)
        # 获得表里的sheet名
        old_sheet_name = old_wb.sheetnames[0]
        # 获得表里的sheet内容
        old_ws = old_wb[old_sheet_name]
        ws = result.create_sheet(sheet_name)

        for row in old_ws.values:
            ws.append(row)
    # 重命名工作表
    result[result.sheetnames[0]].title = '综合报告'
    result[result.sheetnames[1]].title = '业务报告'
    result[result.sheetnames[2]].title = '广告报告'
    result[result.sheetnames[3]].title = '退货报告(FBA)'
    result[result.sheetnames[4]].title = '退货报告(自发货)'
    result.save(filesDir + './临时汇总报告' + processingDir + '.xlsx')

    print('-----***_---\n')
    print('处理完成，临时汇总报告' + processingDir + '.xlsx已生成')
    return


def business_SwitchFile2Xlsx(pyFileDir, outputDir, filesDir, filesArray):
    csvfileName, csvInput, txtfileName, txtInput, xlsxfileName, xlsxInput, xmlfileName, xmlInput = business_getfile(filesArray)
    # --2. 文件格式统一转换为xlsx文件格式
    # --------2.1 csv转xlsx
    datacsv = business_csv2xlsx(filesDir, outputDir, csvfileName, csvInput)
    # --------2.2 txt转xlsx
    theYear, theMonth = business_txt2xlsx(outputDir, filesDir, txtfileName, txtInput)
    # --------2.3 xml转xlsx
    wb = business_xml2xlsx(xmlfileName, filesDir, xmlInput, pyFileDir, outputDir, xlsxInput, xlsxfileName)
    return theYear, theMonth


# --------------------------------------三. 功能实现之覆盖汇总报告模板得到最终汇总报告---------------------------------
def business_GeneratefinalFile(filesDir, processingDir, pyFileDir, theWeek, theYear, theMonth):
    print('-----***_---\n')
    src_wb = openpyxl.load_workbook(filesDir + './临时汇总报告' + processingDir + '.xlsx')
    dst_wb = openpyxl.load_workbook(pyFileDir + './' + '汇总报告_模板.xlsx')

    src_sheets = src_wb.sheetnames
    dst_sheets = dst_wb.sheetnames
    print('源文件工作表和目标文件工作表')
    print(src_sheets)
    print(dst_sheets)
    print('-----***_---\n')

    for i in range(1, len(src_sheets)):
        src_ws = src_wb[src_sheets[i]]
        dst_ws = dst_wb[dst_sheets[i]]
        print('从源' + src_ws.title + '覆盖到目的' + dst_ws.title)
        src_head = 0
        if dst_ws.title == '退货报告(FBA)' or dst_ws.title == '退货报告(自发货)':
            dst_head = 0
        else:
            # 跳过工作表前两行
            dst_head = 2
        for row in range(1, src_ws.max_row + 1):
            src_head = src_head + 1
            dst_head = dst_head + 1
            for i in range(1, src_ws.max_column + 1):
                dst_ws.cell(row=dst_head, column=i).value = src_ws.cell(row=src_head, column=i).value
        # print(src_ws.max_row)
        # print(dst_ws.max_row)
        # print(src_head)
        # print(dst_head)
        if src_ws.max_row < dst_ws.max_row:
            for row in range(dst_head + 1, dst_ws.max_row + 1):
                # print(dst_head)
                dst_head = dst_head + 1
                for i in range(1, src_ws.max_column + 1):
                    dst_ws.cell(row=dst_head, column=i).value = ''
    src_head = 1
    dst_head = 2

    for row in range(1, src_wb['业务报告'].max_row + 1):
        src_head = src_head + 1
        dst_head = dst_head + 1
        dst_wb['综合报告'].cell(row=dst_head, column=1).value = src_wb['业务报告'].cell(row=src_head, column=2).value
        dst_wb['综合报告'].cell(row=dst_head, column=21).value = theYear
        dst_wb['综合报告'].cell(row=dst_head, column=22).value = theMonth
        dst_wb['综合报告'].cell(row=dst_head, column=23).value = theWeek
    if src_wb['业务报告'].max_row < dst_wb['综合报告'].max_row:
        for row in range(dst_head, dst_wb['业务报告'].max_row + 1):
            for col in range(1, dst_wb['业务报告'].max_column + 1):
                dst_wb['综合报告'].cell(row=dst_head, column=col).value = ''
            dst_head = dst_head + 1
    dst_wb.save('汇总报告' + processingDir + '.xlsx')
    src_wb.close()


def business_getTargetFileList(pyFilePath, pyFileDir):
    # --1. 获取输入文件
    file_list = os.listdir(pyFileDir)
    # 先判断是不是文件夹
    tmpList, folderList = [], []
    for file in file_list:
        if os.path.isdir(file) and str(file) != '.idea':
            tmpList.append(file)
    for file in tmpList:
        retFile = '汇总报告' + file + '.xlsx'
        if os.path.exists(retFile):
            print("处理过了，skip")
        else:
            folderList.append(file)
    return folderList


def business_processEntry(processingDir, pyFileDir):
    # ----------------------------------------二. 功能实现之获得转换格式后的临时汇总报告-----------------------------------
    theWeek = processingDir[2:processingDir.find('_DE')]
    filesDir = os.path.join(pyFileDir, processingDir)
    filesArray = os.listdir(processingDir)
    print('All the files are:\n', filesArray)
    print('-----***-----\n')
    outputDir = os.path.join(filesDir, 'outputDir/')
    mkdir(outputDir)
    # --2. 文件格式统一转换为xlsx文件格式
    theYear, theMonth = business_SwitchFile2Xlsx(pyFileDir, outputDir, filesDir, filesArray)
    # --3,合并目录下所有workbook到一个汇总的excel workbook-----------------
    business_combileXlsx(outputDir, filesDir, processingDir)
    # --------------------------------------三. 功能实现之覆盖汇总报告模板得到最终汇总报告---------------------------------
    business_GeneratefinalFile(filesDir, processingDir, pyFileDir, theWeek, theYear, theMonth)


def main():
    # 获取python脚本文件所在绝对路径
    pyFilePath = os.path.abspath(__file__)
    # 获取python脚本文件所在目录
    pyFileDir = os.path.dirname(pyFilePath)
    # ----------------------------------------一. 功能实现之获取待处理文件夹-----------------------------------
    # --1. 获取输入文件
    # --------1.1 当有多个文件夹下数据需要处理时，扫描所有文件，如果未处理就处理
    folderList = business_getTargetFileList(pyFilePath, pyFileDir)
    # 打印所有目标文件
    if len(folderList) is 0:
        print("info: 所有文件夹都处理完了")
        return
    else:
        print("未处理文件夹:",folderList)
    # --------1.1 模式1：批量完所有文件夹，模式2：处理指定输入文件夹（默认无输入是运行模式1）
    processingDir = input("请输入待处理文件夹（无输入会处理完所有文件夹）：")
    if processingDir == "":
        for i in range(len(folderList)):
            # #----------- 1.1.1 获得待处理的目标文件目录
            processingDir = folderList[i]
            business_processEntry(processingDir, pyFileDir)
    else:
        business_processEntry(processingDir, pyFileDir)


if __name__ == "__main__":
    main()
