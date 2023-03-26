#coding=utf-8
#-*- encoding: utf-8 -*-

#---------------------------------------------------一. 导入开源模块-------------------------------------------------
#python2的py文件里面写中文，则必须要添加第1行的声明文件编码的注释
from base64 import encode
import os
# from matplotlib.font_manager import FontProperties  # 获取文件路径需要这个模块
import openpyxl  # 写xlsx文件需要的模块
from openpyxl.styles import PatternFill
import math   #因为广告算法中需要用到e的相关计算，所以因为数学库
import matplotlib.pyplot as plt
import shutil
import warnings

#------------------------------------------------------- 3 Main函数-------------------------------------------------------
if __name__ == '__main__':    
    line = 0 
    KeywordList = []
    ProductList = []
    BidAvgList = []
    flag = 0
    #--------------------------------一. 获得待处理的目标文件目录------------------------
    # 获取python脚本文件所在绝对路径
    pyFilePath = os.path.abspath(__file__)
    # 获取python脚本文件所在目录
    pyFileDir = os.path.dirname(pyFilePath)
    print(pyFileDir, pyFilePath)  
    # 当有多个文件夹下数据需要处理时，可通过此命令选择待处理文件夹
    processingDir = input("please enter your target folder: ")
    print('-----------------------------***************-----------------------\n')
    filesDir = pyFileDir+'./'+processingDir
    # 创建输出文件夹及目录
    outputDir = filesDir +"_proc"
    print(outputDir)  
    # 打印所有目标文件
    filesArray = os.listdir(outputDir)
    print('All the input files are:\n', filesArray)
    print('\n-----------------------------***************-----------------------\n')

    with open(outputDir + '/' +"finalReport.txt","r+", encoding='utf-8') as file:
        for item in file:
            line = line+1
            if item.find("Product_Id") != -1:
                Product_index = item.find("< Product_Id ")   
                Keyword_index = item.find("<Keyword_Id") 
                # print(item[Keyword_index:item.find(">")], item[Product_index:-1])
                KeywordList.append(item[Keyword_index:item.find(">")])
                ProductList.append(item[Product_index:-1])
                retline = line + 6
            if item.find("BidAvg") != -1:                
                # print(item[item.find("BidAvg"):item.find(",")])
                BidAvgList.append(item[item.find("BidAvg"):item.find(",")])
        print(line)
        # print(KeywordList, len(KeywordList))
        # print(ProductList, len(ProductList))
        # print(BidAvgList, len(BidAvgList))
        print(len(KeywordList),len(ProductList),len(BidAvgList))
    file.close()
    file_path = outputDir +'总结.xlsx'
    src_wb = openpyxl.load_workbook(file_path)
    src_ws = src_wb['Sponsored Products Campaigns']
    keyId_column = src_ws['H']
    productId_column = src_ws['I']  
    print('line',line, len(KeywordList))  
    for i in range(len(KeywordList)-1):
        Keyword_Id = (KeywordList[i].split(' '))[1]
        # print(Keyword_Id, Keyword_Id[1])
        Product_Id = (ProductList[i].split(' '))[2]
        # print(Product_Id, Product_Id[2])
        for cell in keyId_column:
            if cell.value == Keyword_Id:
                flag = cell.row
                # print('src_ws',i, flag)
        for cell in productId_column:
            if cell.value == Product_Id:
                flag=cell.row
                # print('src_ws',i, flag)
        # 总结的excel中BidNew在21列
        BidIndex = (BidAvgList[i].split(' '))[1]
        
        Bid = BidIndex.strip()
        print(i, Keyword_Id, Product_Id, BidIndex, type(BidIndex[1]), Bid, BidIndex.strip())
        src_ws.cell(row=flag, column=21).value = float(Bid)
    src_ws.cell(row=1, column=21).value = "BidFinal"
    src_wb.save(file_path)