 #coding=utf-8
 #-*- encoding: utf-8 -*-

#---------------------------------------------------一. 导入开源模块-------------------------------------------------
#python2的py文件里面写中文，则必须要添加第1行的声明文件编码的注释
import os
from matplotlib.font_manager import FontProperties  # 获取文件路径需要这个模块
import openpyxl  # 写xlsx文件需要的模块
import xlrd
import xlwt
from openpyxl.styles import PatternFill
import math   #因为广告算法中需要用到e的相关计算，所以因为数学库
import matplotlib.pyplot as plt
import shutil
import warnings

warnings.filterwarnings('ignore')
#全局变量定义和声明
ACOS_TARGET = 0.36
LOW_CLICkS = 5
UP_BID = 3
DOWN_BID = 0.2
rBidCpc = 1.2   # 因为Bid可能有人会在后台直接修改，所以算新bid的时候用的是旧cpc而不是旧bid。但是cpc往往比bid小一点，所以要乘以一个放大系数。
Price = 12.6
Clicks_Base = 14

#----------------------------------------------三. 功能实现，之计算每个excel里BidNew-------------------------------------------------
# 函数，将文本中的百分数转换为小数
def AD_Perc2Delc(percent):
    percent =(str(percent))
    p_float = float(percent[0:-1])/100
    p_float_2 = round(p_float, 2)
    return p_float_2

# 函数，根据输入转换fileIndex
def AD_ExchangeInput2FileIndex(inputfile):
    fileIndex = 0

    if (inputfile == '14'):
        fileIndex = 0
    elif(inputfile == '30'):
        fileIndex = 1
    elif(inputfile == '60'):
        fileIndex = 2
    return fileIndex

# 函数，从excel读出来的源数据是字符串格式，下面格式转换为可以计算的数字格式
def AD_FmtSrcData(src_head, src_ws):
    AcosStr = src_ws.cell(row=src_head, column=35+1).value
    # 将Acos的字符串百分数转换为小数
    Acos = AD_Perc2Delc(AcosStr)
    Clicks = int(src_ws.cell(row=src_head, column=28+1).value)
    Unit = int(src_ws.cell(row=src_head, column=33+1).value)
    CPC = float(src_ws.cell(row=src_head, column=36+1).value)
    BidOld = float(src_ws.cell(row=src_head, column=21).value)
    return BidOld, Acos, Clicks, Unit, CPC

# 广告算法，根据Acos计算BidNew        
def AD_BidNewAlg(BidOld, Acos, Clicks, Unit, CPC):
    BidNew = 0

    if Acos == 0.0:
        BidNew = 0
    if Acos > 0:
        if Clicks <= LOW_CLICkS:
            Cr_Est = 0.5 * (1/(Clicks))
            Acos = CPC / (Cr_Est * Price)
        x = ACOS_TARGET / Acos
        if x > 1:
            #y =−(1/𝑒) ^ (x−1)+2
            y = (-1 / (math.exp(x-1))) + 2 
        else:  
            # y = 𝑒 ^ (x−1)
            y = math.exp(x-1)
        BidNew = rBidCpc * CPC * y
        # 数据保留2位小数    
    BidNew = round(BidNew,2)
    return BidNew

#打印函数，在文件中把BidNew公式打印出来方便核对计算过程
def AD_BidNewFormularPrint(print_log):
    print("****************************BidNew 计算方法*****************************\n", file=print_log)
    print("<ACOS_TARGET ", ACOS_TARGET,", LOW_CLICkS ", LOW_CLICkS, ", UP_BID ", UP_BID, ", DOWN_BID ", DOWN_BID, ", rBidCpc ", rBidCpc, ", Price ", Price, ", Clicks_Base ", Clicks_Base,">", file=print_log)
    string = \
    "if Acos == 0.0:\n"+\
    "    BidNew = 0"+\
    "if Acos > 0:\n"+\
    "    if Clicks <= LOW_CLICkS:\n"+\
    "        Cr_Est = 0.5 * (1/(Clicks))\n"+\
    "        Acos = CPC / (Cr_Est * Price)\n"+\
    "    x = ACOS_TARGET / Acos\n"+\
    "    if x > 1:\n"+\
    "        y =−(1/𝑒) ^ (x−1)+2\n"+\
    "    else: \n"+\
    "        y = 𝑒 ^ (x−1)\n"+\
    "    1BidNew = rBidCpc * CPC * y\n" +\
    "BidNew = round(BidNew,2)"
    print(string, file=print_log)
    print("\n********************************************************************", file=print_log)

# 函数，BidNew处理模块，处理每行
def AD_DataProc4EachLine(src_head, search_flag, src_ws, print_log):
    # 筛选Product Targeting和Keyword的行
    if search_flag == 'Product Targeting' or search_flag == 'Keyword':            
        BidOld = src_ws.cell(row=src_head, column=21).value
        Keyword_Id = src_ws.cell(row=src_head, column=8).value
        Product_Id = src_ws.cell(row=src_head, column=9).value                               
        # 如果Bid为空则标记红色出来并退出
        if BidOld =="" or BidOld is None:
            fille = PatternFill('solid',fgColor="00FF0000") #标记为红色
            src_ws.cell(row=src_head, column=2).fill = fille
            src_ws.cell(row=src_head, column=3).value = 'BidOld是空'
            print("line ", src_head,"BidOld is None, ", "Keyword_Id:", Keyword_Id, "Product_Id:", Product_Id, file = print_log)
            return
        else:
            # 获得格式化后的源数据
            BidOld, Acos, Clicks, Unit, CPC = AD_FmtSrcData(src_head, src_ws)
            src_ws.cell(row=src_head, column=20).value = BidOld
            # 打印检查获取数据是否正确
            print('\nline', src_head,' < Keyword_Id',Keyword_Id,'> <','Product_Id',Product_Id,'>', file = print_log)
            print(' orig data: ','BidOld',BidOld, 'Acos',Acos, 'Clicks',Clicks, 'Unit',Unit, 'CPC', CPC, file = print_log)
            # 计算BidNew
            BidNew = AD_BidNewAlg(BidOld, Acos, Clicks, Unit, CPC)                        
        # 打印最终结果
        print('line', src_head,' < Keyword_Id',Keyword_Id,'> <','Product_Id',Product_Id,'> <','upper bound',round(UP_BID*BidOld,2),'> <','lower bound',round(DOWN_BID*BidOld,2),'>', file = print_log)
        print('final data: ',"BidOld",BidOld, "BidNew", BidNew, "Acos",Acos, "Clicks",Clicks, "Unit",Unit, "CPC", CPC, file = print_log)
        print('\n', file = print_log)
        return BidNew
    else:
        print("line ", src_head,"is not 'Product Targeting' or 'Keyword', ", "search_flag:", search_flag, '\n', file=print_log)
        return

# 函数，BidNew处理模块，处理每个文件
def AD_BidNewProc(src_head, src_ws, print_log):
    #---------------------------------- 2.1 打开文件---------------------------------------
    src_ws.insert_cols(20, 1)
    src_ws.cell(row=1, column=20).value = 'BidNew'
    # 检查行数，以及下面的列号取的对不对
    #print("src_ws.max_row ", src_ws.max_row)
    if ("Bid" != src_ws.cell(row=1, column=21).value) or\
       ("Keyword Id (Read only)" != src_ws.cell(row=1, column=8).value) or\
       ("Product Targeting Id (Read only)" != src_ws.cell(row=1, column=9).value) or\
       ("Clicks" != src_ws.cell(row=1, column=28+1).value) or\
       ("Units" != src_ws.cell(row=1, column=33+1).value) or\
       ("Acos" != src_ws.cell(row=1, column=35+1).value) or\
       ("CPC" != src_ws.cell(row=1, column=36+1).value):        

        print("列号计算错误")
        print("Bid is ", src_ws.cell(row=1, column=21).value)
        print("Keyword_Id ", src_ws.cell(row=1, column=8).value)
        print("Product_Id ", src_ws.cell(row=1, column=9).value)
        print("Clicks ", src_ws.cell(row=1, column=28+1).value)
        print("Unit ", src_ws.cell(row=1, column=33+1).value)
        print("AcosStr ", src_ws.cell(row=1, column=35+1).value)
        print("CPC ", src_ws.cell(row=1, column=36+1).value)
        return -1
    else:            
        # 在日志文件中打印计算公式
        AD_BidNewFormularPrint(print_log)
        # 循环处理每行
        for row in range(2, src_ws.max_row+1):
            # src_head标记for循环到的当前行数
            src_head = src_head + 1
            search_flag = src_ws.cell(row=src_head, column=2).value
            BidNew = AD_DataProc4EachLine(src_head, search_flag, src_ws, print_log)
            src_ws.cell(row=src_head, column=20).value = BidNew
        return 0
#--------------------------------------四. 功能实现之计算14天excel里BidAvg---------------------------------
print('\n-----***-----\n')
print('功能实现之计算14天excel里BidAvg\n')
print('\n-----***-----\n')

# 广告算法，根据Click和BidNew计算BidAvg   
def AD_BidAvgAlg(Clicks_14, Clicks_30, Clicks_60, BidNew_14, BidNew_30, BidNew_60, CPC_14, print_log):
    #timeCoe14\timeCoe30\timeCoe60是时间权重
    weight14 = 0
    weight30 = 0
    weight60 = 0
    weightSum = 0
    if (BidNew_14+BidNew_30+BidNew_60) == 0:
        BidAvg = 999
        fille = PatternFill('solid',fgColor="001874CD") #蓝色
    else:
        if Clicks_14 <= 100:
            timeCoe14 = 0.5
            timeCoe30 = 0.3
            timeCoe60 = 0.2 
        elif Clicks_14 > 100 and Clicks_14 <= 250:
            timeCoe14 = 0.7
            timeCoe30 = 0.2
            timeCoe60 = 0.1
        elif Clicks_14 > 250 and Clicks_14 <= 500:
            timeCoe14 = 0.8
            timeCoe30 = 0.15
            timeCoe60 = 0.05
        else:
            timeCoe14 = 0.9
            timeCoe30 = 0.1
            timeCoe60 = 0

        if (BidNew_14!=0):
            weight14 = timeCoe14 * Clicks_14
        if (BidNew_30!=0):
            weight30 = timeCoe30 * Clicks_30
        if (BidNew_60!=0):
            weight60 = timeCoe60 * Clicks_60
        weightSum = weight14 + weight30 + weight60
        BidAvg = float(BidNew_14) * weight14 / weightSum+ float(BidNew_30) * weight30 / weightSum + float(BidNew_60) * weight60 / weightSum

        if BidAvg > CPC_14:
            if Clicks_14 > Clicks_Base:
                fille = PatternFill('solid',fgColor="00008000") #深绿色       
            else:
                fille = PatternFill('solid',fgColor="00CCFFCC") #浅绿色
        if BidAvg < CPC_14 and BidAvg > 0:
            if Clicks_14 > Clicks_Base:
                fille = PatternFill('solid',fgColor="00FF6600") #深橙色       
            else:
                fille = PatternFill('solid',fgColor="00FFCC99") #浅橙色    
    print("weightSum ", weightSum, "weight14 ", round(weight14,3) ,"weight30 ", round(weight30,3), "weight60 ", round(weight60,3), file=print_log)
    return BidAvg, fille

#打印函数，在文件中把BidNew公式打印出来方便核对计算过程
def AD_BidAvgFormularPrint(print_log):
    print("****************************BidAvg计算方法*****************************", file=print_log)
    string = '''
    BidAvg = float(BidNew_14) * weight14 / weightSum+ float(BidNew_30) * weight30 / weightSum + float(BidNew_60) * weight60 / weightSum
    '''
    print(string, file=print_log)
    print("********************************************************************", file=print_log)

# 函数，BidAvg处理模块，处理每个行
def AD_BidAvgProc(src_head, search_flag, src_ws_14, src_ws_30, src_ws_60, BidChgList, BidOldList, BidAvgList, print_log):
    if search_flag == 'Product Targeting' or search_flag == 'Keyword':            
        BidOld_14 = src_ws_14.cell(row=src_head, column=22).value
        Keyword_Id = src_ws_14.cell(row=src_head, column=8).value
        Product_Id = src_ws_14.cell(row=src_head, column=9).value
        # 如果Bid为空则退出
        if BidOld_14 =="" or BidOld_14 is None:
            print("line ", src_head,"BidOld_14 is None, ", "Keyword_Id:", Keyword_Id, "Product_Id:", Product_Id, file = print_log)
            return
        else:
            Clicks_14 = int(src_ws_14.cell(row=src_head, column=30).value)
            BidNew_14 = src_ws_14.cell(row=src_head, column=21).value
            # print('Keyword_Id', Keyword_Id,'Product_Id',Product_Id,'Clicks_14',Clicks_14, 'BidOld_14',BidOld_14,'BidNew_14',BidNew_14)
            if Keyword_Id is not None:
                keyId_column = src_ws_30['H']
                for cell in keyId_column:
                    if cell.value == Keyword_Id:
                        flag = cell.row
                        # print('src_ws_30',head, flag)
                        Clicks_30 = int(src_ws_30.cell(row=flag, column=29).value)
                        BidNew_30 = src_ws_30.cell(row=flag, column=20).value
                        # print('Clicks_30',Clicks_30,'BidNew_30',BidNew_30)
                        break
                keyId_column = src_ws_60['H']
                for cell in keyId_column:
                    if cell.value == Keyword_Id:
                        flag=cell.row
                        # print('src_ws_60',head, flag)
                        Clicks_60 = int(src_ws_60.cell(row=flag, column=29).value)
                        BidNew_60 = src_ws_60.cell(row=flag, column=20).value
                        # print('Clicks_60',Clicks_60,'BidNew_60',BidNew_60)
                        break
            if Product_Id is not None:
                productId_column = src_ws_30['I']
                for cell in productId_column:
                    if cell.value == Product_Id:
                        flag=cell.row
                        # print('src_ws_30',head, flag)
                        Clicks_30 = int(src_ws_30.cell(row=flag, column=29).value)
                        BidNew_30 = src_ws_30.cell(row=flag, column=20).value
                        # print('Clicks_30',Clicks_30,'BidNew_30',BidNew_30)
                        break
                productId_column = src_ws_60['I']
                for cell in productId_column:
                    if cell.value == Product_Id:
                        flag=cell.row
                        # print('src_ws_60',head, flag)
                        Clicks_60 = int(src_ws_60.cell(row=flag, column=29).value)
                        BidNew_60 = src_ws_60.cell(row=flag, column=20).value
                        # print('Clicks_60',Clicks_60,'BidNew_60',BidNew_60)
                        break

            ClicksSum = Clicks_14 + Clicks_30 + Clicks_60
            if BidNew_60 is None:
                print("\nline ", src_head, '<Keyword_Id',Keyword_Id,'> <','Product_Id',Product_Id,'>', file=print_log)
                print('BidNew_14',BidNew_14,'BidNew_30',BidNew_30,'BidNew_60',BidNew_60, '\n',file=print_log)
                return
            if ClicksSum == 0:
                print("\nline ", src_head, '<Keyword_Id',Keyword_Id,'> <','Product_Id',Product_Id,'>', file=print_log)
                print('ClicksSum is 0\n', file=print_log)
                return

            print("\nline ", src_head, '<Keyword_Id',Keyword_Id,'> <','Product_Id',Product_Id,'>', file = print_log)
            print('ClicksSum',ClicksSum,' ','Clicks_14',Clicks_14,' ','Clicks_30',Clicks_30,' ','Clicks_60',Clicks_60, file = print_log)
            print('BidNew_14',BidNew_14, 'BidNew_30',BidNew_30,'BidNew_60',BidNew_60, file = print_log)

            CPC_14 = float(src_ws_14.cell(row=src_head, column=38).value)
            BidAvg, fille = AD_BidAvgAlg(Clicks_14, Clicks_30, Clicks_60, BidNew_14, BidNew_30, BidNew_60, CPC_14, print_log)
            BidAvgList[src_head] = BidAvg
            if BidAvg == 999:                
                src_ws_14.cell(row=src_head, column=3).value = 'BidAvg手动改'
                src_ws_14.cell(row=src_head, column=20).value = '--'
            else:
                BidChgList[src_head] = BidAvg - CPC_14
                BidOldList[src_head] = src_ws_14.cell(row=src_head,column=22).value                
                src_ws_14.cell(row=src_head, column=3).value = 'BidAvg手动改'
                src_ws_14.cell(row=src_head, column=20).value = round(BidAvg,2)        
            src_ws_14.cell(row=src_head, column=2).fill = fille
            print('final: BidAvg',round(BidAvg,2), '\n',file=print_log)
            return
    else:
        print("line ", src_head,"is not 'Product Targeting' or 'Keyword', ", "search_flag:", search_flag, file = print_log)
        return

# 函数，画图模块
def AD_PlotResult(row, BidOldList, BidAvgList, BidChgList):
    listlen = len(BidChgList)  
    list = [i for i in range(listlen)]
    plt.title('BigAvg-CPC ChangeFlow')
    # 设置坐标轴
    plt.xlabel('excel中行号')
    plt.ylabel('BidAvg-CPC_14')
    # 设置坐标轴范围
    # plt.xticks([0,row])
    plt.xlim(0,row)
    plt.ylim(-1,1)
    # plt.yticks([-1,6])
    # 画两条虚线
    plt.grid()
    plt.hlines(0.5,0,row, colors='g', linestyle='--')
    plt.hlines(0,0,row, colors='r', linestyle='--')
    plt.hlines(-0.5,0,row, colors='g',linestyle='--')
    # plt.scatter(x= list, y=BidOldList, marker='X', c='r', s=20)
    # plt.scatter(x= list, y=BidAvgList, marker='o', c='g', s=20)
    plt.scatter(x= list, y=BidChgList, marker='*', c='y', s=20)
    for i in range(listlen):
        if(BidChgList[i]>0):
            plt.annotate(i, xy=(list[i], BidChgList[i]),xytext=(list[i]+0.1,BidChgList[i]+0.1))
    plt.savefig(outputDir+ '/' +'Bid变化图.png')
    plt.show()

#------------------------------------------------------- 3 Main函数-------------------------------------------------------

if __name__ == '__main__':    

    #--------------------------------一. 获得待处理的目标文件目录------------------------
    # 获取python脚本文件所在绝对路径
    pyFilePath = os.path.abspath(__file__)
    # 获取python脚本文件所在目录
    pyFileDir = os.path.dirname(pyFilePath)
    print(pyFileDir, pyFilePath)  
    # 当有多个文件夹下数据需要处理时，可通过此命令选择待处理文件夹
    processingDir = input("please enter your target folder: ")
    print('-----***-----\n')
    filesDir = pyFileDir+'./'+processingDir
    # 打印所有目标文件
    filesArray = os.listdir(processingDir)
    print('All the input files are:\n', filesArray)
    print('\n-----***-----\n')
    # 创建输出文件夹及目录
    outputDir = filesDir +"_proc"  
    os.makedirs(outputDir)
    #将原来已有的xlsx文件复制到output文件夹
    for file in filesArray:
        shutil.copy(filesDir+'/'+file, outputDir+'/'+file)
                        
    print('输出文件目录:', outputDir)
    # 获取输出目录下所有的excel表
    file_list = os.listdir(outputDir)
    print('All the output files are:\n', filesArray)         
    print('-----***-----\n')

    #---------------------------二. 功能实现, 之计算每个excel里BidNew---------------------
    with open(outputDir + '/' +"calBidNewlog.txt","w", encoding='utf-8') as print_log:
        for i in range(3):
            # inputfile = input("input 14 or 30 or 60: ")
            # fileIndex = AD_ExchangeInput2FileIndex(inputfile)
            fileIndex = i
            file_path = outputDir + '/' + file_list[fileIndex]
            src_head = 1
            # 打开待处理文件
            src_wb = openpyxl.load_workbook(file_path)
            src_ws = src_wb['Sponsored Products Campaigns']
            # 计算每个文件里所有BidNew
            ret = AD_BidNewProc(src_head, src_ws, print_log)
            if ret != 0:
                print("error!!!")
                break
            # 打印处理进度
            print((i+1)*5,'%')
            # 保存文件
            src_wb.save(file_path)
            print(file_path ," 计算完成")
            if (i==0):
                src_wb_14 = src_wb
                src_ws_14 = src_ws
            if (i==1):
                src_wb_30 = src_wb
                src_ws_30 = src_ws
            if (i==2):
                src_wb_60 = src_wb
                src_ws_60 = src_ws
        print('-------------all file cal BidNew finished-----------')
        print('\n----------------------end------------------------\n', file=print_log)
    print_log.close
    #---------------------------三. 功能实现, 之计算14天excel里BidAvg----------------------
    # src_wb_14= openpyxl.load_workbook(outputDir + '/' +'14天.xlsx')
    # src_wb_30= openpyxl.load_workbook(outputDir + '/' +'30天.xlsx')
    # src_wb_60= openpyxl.load_workbook(outputDir + '/' +'60天.xlsx')        
    # src_ws_14 = src_wb_14['Sponsored Products Campaigns']
    # src_ws_30 = src_wb_30['Sponsored Products Campaigns']
    # src_ws_60 = src_wb_60['Sponsored Products Campaigns']
    # 插入BidAvg的列
    src_ws_14.insert_cols(20,1)
    src_ws_14.cell(row=1, column=20).value ='BidAvg'
    
    BidChgList = [0 for i in range(src_ws_14.max_row+2)]
    BidAvgList = [0 for i in range(src_ws_14.max_row+2)]
    BidOldList = [0 for i in range(src_ws_14.max_row+2)]
    file_path = outputDir + '/' +'总结.xlsx'
    print('20%')
    #记录打印日志
    with open(outputDir + '/' +"calBidAvglog.txt","w") as print_log:
        AD_BidAvgFormularPrint(print_log)
        src_head = 2
        for row in range(2, src_ws_14.max_row+1):
            src_head=src_head+1
            # 显示处理进度
            if (src_head%100 == 0):
                print(round(20 +100*src_head/src_ws_14.max_row) %100,'%')
            # 计算所有BidAvg
            search_flag = src_ws_14.cell(row=src_head, column=2).value
            AD_BidAvgProc(src_head, search_flag, src_ws_14, src_ws_30, src_ws_60, BidChgList, BidOldList, BidAvgList, print_log)
       
        print('\n----------------------end------------------------\n', file=print_log)
    print_log.close
    print('100%')
    src_wb_14.save(file_path)
    print('-------------all file cal BidAvg finished-----------\n')
    
    AD_PlotResult(src_ws_14.max_row+2, BidOldList, BidAvgList, BidChgList)



